from flask import Flask, render_template, request, jsonify, send_file
from ultralytics import YOLO
import cv2
import os
import uuid
import json
from datetime import datetime
from fpdf import FPDF
import base64
import io
import zipfile
import tempfile
import pymssql
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv
load_dotenv()

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))


AZURE_CONNECTION_STRING = os.getenv('AZURE_CONNECTION_STRING')
SQL_SERVER   = os.getenv('SQL_SERVER')
SQL_DATABASE = os.getenv('SQL_DATABASE')
SQL_USER     = os.getenv('SQL_USER')
SQL_PASSWORD = os.getenv('SQL_PASSWORD')
AZURE_CONTAINER         = "batch-images"


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

# ── paths ──────────────────────────────────────────────────────────────
MODEL_PATH  = r'C:\Users\nia.insyirah\models\yolo_mopad\weights\best.pt'
UPLOAD_DIR  = os.path.join(os.path.dirname(__file__), 'uploads')
RESULTS_DIR = os.path.join(os.path.dirname(__file__), 'results')
os.makedirs(UPLOAD_DIR,  exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)

model = YOLO(MODEL_PATH)

# ── Azure clients ──────────────────────────────────────────────────────
try:
    blob_service = BlobServiceClient.from_connection_string(AZURE_CONNECTION_STRING)
except Exception as e:
    print(f"Blob Storage connection failed: {e}")
    blob_service = None


# ── DB helpers ─────────────────────────────────────────────────────────
def get_db_connection():
    return pymssql.connect(
        server=SQL_SERVER,
        database=SQL_DATABASE,
        user=SQL_USER,
        password=SQL_PASSWORD
    )


def setup_database():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='batch_results' AND xtype='U')
            CREATE TABLE batch_results (
                id INT IDENTITY(1,1) PRIMARY KEY,
                batch_uid NVARCHAR(50),
                batch_name NVARCHAR(200),
                timestamp DATETIME,
                total_images INT,
                total_trees INT,
                healthy INT,
                small_trees INT,
                yellow INT,
                dead INT,
                healthy_pct FLOAT
            )
        ''')
        conn.commit()
        conn.close()
        print("Database setup complete")
    except Exception as e:
        print(f"Database setup failed: {e}")


def save_batch_to_db(batch_uid, batch_name, total_images, total_trees, counts):
    try:
        healthy_pct = round((counts['Healthy'] / total_trees * 100), 1) if total_trees > 0 else 0
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO batch_results
            (batch_uid, batch_name, timestamp, total_images, total_trees,
             healthy, small_trees, yellow, dead, healthy_pct)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            batch_uid, batch_name, datetime.now(),
            total_images, total_trees,
            counts['Healthy'], counts['Small'], counts['Yellow'], counts['Dead'],
            healthy_pct
        ))
        conn.commit()
        conn.close()
        print(f"Batch {batch_name} saved to Azure SQL")
    except Exception as e:
        print(f"Failed to save to Azure SQL: {e}")


# ── inference helpers ──────────────────────────────────────────────────
def img_to_b64(arr_bgr):
    _, buf = cv2.imencode('.jpg', arr_bgr, [cv2.IMWRITE_JPEG_QUALITY, 92])
    return base64.b64encode(buf).decode()


def run_inference(img_path):
    results   = model.predict(source=img_path, imgsz=640, conf=0.25, verbose=False)
    result    = results[0]
    annotated = result.plot()

    counts = {'Dead': 0, 'Healthy': 0, 'Small': 0, 'Yellow': 0}
    confs  = {'Dead': [], 'Healthy': [], 'Small': [], 'Yellow': []}

    for box in result.boxes:
        cls_id   = int(box.cls[0])
        cls_name = result.names[cls_id]
        conf     = float(box.conf[0])
        if cls_name in counts:
            counts[cls_name] += 1
            confs[cls_name].append(conf)

    avg_conf = {k: (sum(v)/len(v) if v else 0) for k, v in confs.items()}
    total    = sum(counts.values())
    return annotated, counts, avg_conf, total


def upload_to_blob(file_path, blob_name):
    if blob_service is None:
        return False
    try:
        blob_client = blob_service.get_blob_client(
            container=AZURE_CONTAINER, blob=blob_name)
        with open(file_path, 'rb') as f:
            blob_client.upload_blob(f, overwrite=True)
        return True
    except Exception as e:
        print(f"Blob upload failed: {e}")
        return False


def process_one_image(file_obj, filename, batch_name=None):
    """
    Save, run inference, upload to blob, return result dict.
    Used by both /predict and /predict_batch.
    """
    ext      = os.path.splitext(filename)[1].lower()
    uid      = uuid.uuid4().hex
    src_path = os.path.join(UPLOAD_DIR, f'{uid}{ext}')
    file_obj.save(src_path)

    annotated, counts, avg_conf, total = run_inference(src_path)

    res_path = os.path.join(RESULTS_DIR, f'{uid}_result.jpg')
    cv2.imwrite(res_path, annotated)

    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta = {
        'uid': uid, 'filename': filename,
        'timestamp': ts, 'counts': counts,
        'avg_conf': avg_conf, 'total': total,
    }
    with open(os.path.join(RESULTS_DIR, f'{uid}_meta.json'), 'w') as f:
        json.dump(meta, f)

    if batch_name:
        upload_to_blob(src_path, f"{batch_name}/{filename}")

    return {
        'uid':       uid,
        'filename':  filename,
        'timestamp': ts,
        'counts':    counts,
        'avg_conf':  {k: round(v*100, 1) for k, v in avg_conf.items()},
        'total':     total,
        'image_b64': img_to_b64(annotated),
    }


# ── routes ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# 1. SINGLE IMAGE ────────────────────────────────────────────────────────
@app.route('/predict', methods=['POST'])
def predict():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400
    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    try:
        result = process_one_image(file, file.filename)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 2. BATCH — MULTIPLE IMAGES (matches index.html frontend) ───────────────
@app.route('/predict_batch', methods=['POST'])
def predict_batch():
    files = request.files.getlist('images')
    if not files or files[0].filename == '':
        return jsonify({'error': 'No images uploaded'}), 400

    batch_name = request.form.get(
        'batch_name',
        f'Batch_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    )
    batch_uid = uuid.uuid4().hex

    results = []
    errors  = []

    for file in files:
        try:
            result = process_one_image(file, file.filename, batch_name)
            results.append(result)
        except Exception as e:
            print(f"Failed on {file.filename}: {e}")
            errors.append({'filename': file.filename, 'error': str(e)})

    if not results:
        return jsonify({'error': 'All images failed to process'}), 500

    # summary totals
    total_counts = {'Dead': 0, 'Healthy': 0, 'Small': 0, 'Yellow': 0}
    total_trees  = 0
    for r in results:
        for cls in total_counts:
            total_counts[cls] += r['counts'][cls]
        total_trees += r['total']

    # save summary to Azure SQL
    save_batch_to_db(batch_uid, batch_name, len(results), total_trees, total_counts)

    healthy_pct = round(total_counts['Healthy'] / total_trees * 100, 1) if total_trees > 0 else 0

    return jsonify({
        'batch_name':   batch_name,
        'total_images': len(results),
        'total_trees':  total_trees,
        'counts':       total_counts,
        'healthy_pct':  healthy_pct,
        'results':      results,   # ← index.html reads data.results
        'errors':       errors,    # ← index.html reads data.errors
        'message':      f'Processed {len(results)} images. Saved to Azure.'
    })


# 3. BATCH — ZIP FILE ────────────────────────────────────────────────────
@app.route('/batch', methods=['POST'])
def batch_zip():
    if 'zipfile' not in request.files:
        return jsonify({'error': 'No ZIP file uploaded'}), 400

    zip_file   = request.files['zipfile']
    batch_name = request.form.get(
        'batch_name',
        f'Batch_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    )
    batch_uid = uuid.uuid4().hex

    temp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_file, 'r') as z:
            z.extractall(temp_dir)
    except Exception as e:
        return jsonify({'error': f'Invalid ZIP file: {str(e)}'}), 400

    valid_ext   = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp'}
    image_files = []
    for root, dirs, files in os.walk(temp_dir):
        for f in files:
            if os.path.splitext(f)[1].lower() in valid_ext:
                image_files.append(os.path.join(root, f))

    if not image_files:
        return jsonify({'error': 'No valid images found in ZIP'}), 400

    total_counts = {'Dead': 0, 'Healthy': 0, 'Small': 0, 'Yellow': 0}
    total_trees  = 0
    processed    = 0
    failed       = 0

    for img_path in image_files:
        try:
            blob_name = f"{batch_name}/{os.path.basename(img_path)}"
            upload_to_blob(img_path, blob_name)
            _, counts, _, total = run_inference(img_path)
            for cls in total_counts:
                total_counts[cls] += counts[cls]
            total_trees += total
            processed   += 1
        except Exception as e:
            print(f"Failed on {img_path}: {e}")
            failed += 1

    if processed == 0:
        return jsonify({'error': 'All images failed to process'}), 500

    save_batch_to_db(batch_uid, batch_name, processed, total_trees, total_counts)
    healthy_pct = round(total_counts['Healthy'] / total_trees * 100, 1) if total_trees > 0 else 0

    return jsonify({
        'batch_name':   batch_name,
        'total_images': processed,
        'total_trees':  total_trees,
        'counts':       total_counts,
        'healthy_pct':  healthy_pct,
        'failed':       failed,
        'timestamp':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'message':      f'Processed {processed} images. Saved to Azure SQL.'
    })


# 4. HISTORY ─────────────────────────────────────────────────────────────
@app.route('/history', methods=['GET'])
def get_history():
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT batch_name, timestamp, total_images, total_trees,
                   healthy, small_trees, yellow, dead, healthy_pct
            FROM batch_results ORDER BY timestamp DESC
        ''')
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            'batch_name':   r[0], 'timestamp':    str(r[1]),
            'total_images': r[2], 'total_trees':  r[3],
            'healthy':      r[4], 'small':        r[5],
            'yellow':       r[6], 'dead':         r[7],
            'healthy_pct':  r[8],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 5. PDF REPORT ───────────────────────────────────────────────────────────
@app.route('/report/<uid>')
def download_report(uid):
    meta_path = os.path.join(RESULTS_DIR, f'{uid}_meta.json')
    img_path  = os.path.join(RESULTS_DIR, f'{uid}_result.jpg')
    if not os.path.exists(meta_path):
        return 'Report not found', 404

    with open(meta_path) as f:
        meta = json.load(f)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_fill_color(22, 101, 52)
    pdf.rect(0, 0, 210, 30, 'F')
    pdf.set_font('Helvetica', 'B', 20)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, '', ln=True)
    pdf.cell(0, 12, 'MOPAD Tree Health Report', align='C', ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 10)
    pdf.ln(8)
    pdf.cell(0, 6, f"File: {meta['filename']}", ln=True)
    pdf.cell(0, 6, f"Analysis Date: {meta['timestamp']}", ln=True)
    pdf.cell(0, 6, f"Total Trees Detected: {meta['total']}", ln=True)
    pdf.ln(4)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_fill_color(220, 252, 231)
    pdf.cell(60, 8, 'Class',      border=1, fill=True)
    pdf.cell(40, 8, 'Count',      border=1, fill=True)
    pdf.cell(40, 8, 'Percentage', border=1, fill=True)
    pdf.cell(40, 8, 'Avg Conf %', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 10)
    total = meta['total'] or 1
    for cls in ['Dead', 'Healthy', 'Small', 'Yellow']:
        cnt  = meta['counts'].get(cls, 0)
        pct  = round(cnt / total * 100, 1)
        conf = round(meta['avg_conf'].get(cls, 0) * 100, 1)
        pdf.cell(60, 7, cls,        border=1)
        pdf.cell(40, 7, str(cnt),   border=1)
        pdf.cell(40, 7, f'{pct}%',  border=1)
        pdf.cell(40, 7, f'{conf}%', border=1, ln=True)
    if os.path.exists(img_path):
        pdf.ln(6)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 8, 'Annotated Image:', ln=True)
        pdf.image(img_path, x=10, w=190)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'tree_report_{uid[:8]}.pdf')

# 6. EXCEL EXPORT ─────────────────────────────────────────────────────────
@app.route('/report_batch_excel', methods=['POST'])
def report_batch_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        data    = request.get_json()
        results = data.get('results', [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detection Results"

        # header style
        green_fill = PatternFill("solid", fgColor="166534")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ['No', 'Filename', 'Timestamp', 'Total', 'Healthy', 'Small', 'Yellow', 'Dead', 'Healthy %']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = green_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # data rows
        for i, r in enumerate(results, 1):
            total = r.get('total', 0) or 1
            healthy_pct = round(r['counts']['Healthy'] / total * 100, 1)
            ws.append([
                i,
                r.get('filename', ''),
                r.get('timestamp', ''),
                r.get('total', 0),
                r['counts']['Healthy'],
                r['counts']['Small'],
                r['counts']['Yellow'],
                r['counts']['Dead'],
                f"{healthy_pct}%"
            ])

        # auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=f'mopad_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    setup_database()
    app.run(debug=True, port=5000)
